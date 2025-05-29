import soundfile as sf
import os
import openpyxl
import wave
import shutil

from openpyxl import Workbook

workbook = Workbook()
print(workbook.sheetnames)
workbook.save(filename="infofile.xlsx")
wb = openpyxl.load_workbook("infofile.xlsx")
sheet = wb['Sheet']
pathoffolder = 'positive'
files = os.listdir(pathoffolder)
columnvalue = 1
for k in range(1, sheet.max_row + 1):
    if str(sheet.cell(row=k, column=columnvalue).value) == 'None':
        cell = sheet.cell(row=k, column=columnvalue)
        cell.value = "filename"

        columnvalue+=1

        break
for k in range(1, sheet.max_row + 1):
    if str(sheet.cell(row=k, column=columnvalue).value) == 'None':
        cell = sheet.cell(row=k, column=columnvalue)
        cell.value = "subtype"
        columnvalue+=1

        break
for k in range(1, sheet.max_row + 1):
    if str(sheet.cell(row=k, column=columnvalue).value) == 'None':
        cell = sheet.cell(row=k, column=columnvalue)
        cell.value = "format"
        columnvalue+=1

        break
for k in range(1, sheet.max_row + 1):
    if str(sheet.cell(row=k, column=columnvalue).value) == 'None':
        cell = sheet.cell(row=k, column=columnvalue)
        cell.value = "channels"
        columnvalue+=1

        break
for k in range(1, sheet.max_row + 1):
    if str(sheet.cell(row=k, column=columnvalue).value) == 'None':
        cell = sheet.cell(row=k, column=columnvalue)
        cell.value = "format_info"
        columnvalue+=1

        break
for k in range(1, sheet.max_row + 1):
    if str(sheet.cell(row=k, column=columnvalue).value) == 'None':
        cell = sheet.cell(row=k, column=columnvalue)
        cell.value = "frames"
        columnvalue+=1

        break
for k in range(1, sheet.max_row + 1):
    if str(sheet.cell(row=k, column=columnvalue).value) == 'None':
        cell = sheet.cell(row=k, column=columnvalue)
        cell.value = "samplerate"
        columnvalue+=1

        break
for k in range(1, sheet.max_row + 1):
    if str(sheet.cell(row=k, column=columnvalue).value) == 'None':
        cell = sheet.cell(row=k, column=columnvalue)
        cell.value = "mode"
        columnvalue+=1

        break
for k in range(1, sheet.max_row + 1):
    if str(sheet.cell(row=k, column=columnvalue).value) == 'None':
        cell = sheet.cell(row=k, column=columnvalue)
        cell.value = "subtype_info"
        columnvalue+=1

        break
for k in range(1, sheet.max_row + 1):
    if str(sheet.cell(row=k, column=columnvalue).value) == 'None':
        cell = sheet.cell(row=k, column=columnvalue)
        cell.value = "sections"
        columnvalue+=1

        break

for i in files:
    i = pathoffolder+"/"+i
    file = i

    ob = sf.SoundFile(file)

    if ob.channels == 2:
        os.remove(i)
        continue
    selected_file = wave.open(file)
    frames = selected_file.getnframes()
    rate = selected_file.getframerate()
    duration_of_selected = frames / float(rate)

    if duration_of_selected < 3:
        shutil.move(file,"positive_dustbin")
        continue

    columnvalue = 1
    for k in range(1, sheet.max_row + 2):
        if str(sheet.cell(row=k, column=columnvalue).value) == 'None':
            cell = sheet.cell(row=k, column=columnvalue)
            cell.value = ob.name
            columnvalue+=1

            break
    for k in range(1, sheet.max_row + 2):
        if str(sheet.cell(row=k, column=columnvalue).value) == 'None':
            cell = sheet.cell(row=k, column=columnvalue)
            cell.value = ob.subtype
            columnvalue+=1

            break
    for k in range(1, sheet.max_row + 2):
        if str(sheet.cell(row=k, column=columnvalue).value) == 'None':
            cell = sheet.cell(row=k, column=columnvalue)
            cell.value = ob.format
            columnvalue+=1

            break
    for k in range(1, sheet.max_row + 2):
        if str(sheet.cell(row=k, column=columnvalue).value) == 'None':
            cell = sheet.cell(row=k, column=columnvalue)
            cell.value = ob.channels
            columnvalue+=1

            break
    for k in range(1, sheet.max_row + 2):
        if str(sheet.cell(row=k, column=columnvalue).value) == 'None':
            cell = sheet.cell(row=k, column=columnvalue)
            cell.value = ob.format_info
            columnvalue+=1

            break
    for k in range(1, sheet.max_row +2 ):
        if str(sheet.cell(row=k, column=columnvalue).value) == 'None':
            cell = sheet.cell(row=k, column=columnvalue)
            cell.value = ob.frames
            columnvalue+=1

            break
    for k in range(1, sheet.max_row + 2):
        if str(sheet.cell(row=k, column=columnvalue).value) == 'None':
            cell = sheet.cell(row=k, column=columnvalue)
            cell.value = ob.samplerate
            columnvalue+=1

            break
    for k in range(1, sheet.max_row + 2):
        if str(sheet.cell(row=k, column=columnvalue).value) == 'None':
            cell = sheet.cell(row=k, column=columnvalue)
            cell.value = ob.mode
            columnvalue+=1

            break
    for k in range(1, sheet.max_row + 2):
        if str(sheet.cell(row=k, column=columnvalue).value) == 'None':
            cell = sheet.cell(row=k, column=columnvalue)
            cell.value = ob.subtype_info
            columnvalue+=1

            break
    for k in range(1, sheet.max_row + 2):
        if str(sheet.cell(row=k, column=columnvalue).value) == 'None':
            cell = sheet.cell(row=k, column=columnvalue)
            cell.value = ob.sections
            columnvalue+=1

            break

wb.save("infofile.xlsx")